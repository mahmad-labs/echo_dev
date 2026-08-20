from __future__ import annotations

from pathlib import Path
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


class DocumentExtractor:
    def extract(self, path):
        file_path = Path(path)
        suffix = file_path.suffix.lower()
        if suffix in {'.txt', '.md', '.csv', '.json', '.xml', '.html', '.css', '.js', '.py'}:
            return file_path.read_text(encoding='utf-8', errors='replace')
        if suffix == '.pdf':
            return '\n'.join(page.extract_text() or '' for page in PdfReader(file_path).pages)
        if suffix == '.docx':
            return '\n'.join(paragraph.text for paragraph in DocxDocument(file_path).paragraphs)
        if suffix in {'.xlsx', '.xlsm'}:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for worksheet in workbook.worksheets:
                rows.append(f'# {worksheet.title}')
                rows.extend(
                    '\t'.join('' if value is None else str(value) for value in row)
                    for row in worksheet.iter_rows(values_only=True)
                )
            return '\n'.join(rows)
        raise ValueError(f'Unsupported document type: {suffix}')
